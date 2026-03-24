"""
generate_excel.py
Generates a styled Excel file matching Registre_des_traitement_vide.xlsm
Called from the browser tool via: python3 generate_excel.py <json_input> <output_path>
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ─── STYLE CONSTANTS (from original XLSM) ────────────────────────────────────
GREEN_HEADER   = "FF267959"   # column header bg
WHITE_FONT     = "FFFFFFFF"   # white text
DARK_RED_FONT  = "FFC00000"   # title text
NAVY_FILL      = "FF002060"   # section header alt (not used, keeping green)
SECTION_BG     = "FFE2EFDA"   # light green for section title rows
DATA_BG_ALT    = "FFEBF5F0"   # alternating row tint

thin  = Side(style="thin",  color="FF000000")
thick = Side(style="medium", color="FF000000")

def hdr_font(sz=10):   return Font(name="Calibri", bold=True,  size=sz, color=WHITE_FONT)
def data_font(sz=10):  return Font(name="Calibri", bold=False, size=sz)
def title_font(sz=14): return Font(name="Calibri", bold=True,  size=sz)
def big_font(sz=22, color=DARK_RED_FONT): return Font(name="Calibri", bold=True, size=sz, color=color)

def green_fill(): return PatternFill("solid", fgColor=GREEN_HEADER)
def light_fill(): return PatternFill("solid", fgColor=SECTION_BG)
def alt_fill():   return PatternFill("solid", fgColor=DATA_BG_ALT)
def no_fill():    return PatternFill("none")

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def all_borders():
    return Border(left=thin, right=thin, top=thin, bottom=thin)
def bottom_only():
    return Border(bottom=thin)
def top_bottom():
    return Border(top=thin, bottom=thin)

def style_header_cell(cell, value):
    cell.value = value
    cell.font = hdr_font()
    cell.fill = green_fill()
    cell.alignment = center()
    cell.border = all_borders()

def style_data_cell(cell, value, alt=False):
    cell.value = value
    cell.font = data_font()
    cell.fill = alt_fill() if alt else no_fill()
    cell.alignment = center()
    cell.border = all_borders()

def style_section_title(cell, value):
    cell.value = value
    cell.font = title_font(14)
    cell.fill = light_fill()
    cell.alignment = left_align()
    cell.border = bottom_only()

def set_col_widths(ws, widths):
    """widths: list of (col_letter, width)"""
    for col, w in widths:
        ws.column_dimensions[col].width = w

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


# ─── BUILD SHEET ─────────────────────────────────────────────────────────────
def build_sheet(wb, sheet_name, d, label):
    ws = wb.create_sheet(title=sheet_name)

    # Column widths matching original
    col_widths = [
        ('A',35),('B',33),('C',22),('D',30),('E',30),('F',26),
        ('G',22),('H',24),('I',22),('J',22),('K',20),('L',22),
        ('M',28),('N',20),('O',22),('P',24),('Q',30),('R',20),('S',20)
    ]
    set_col_widths(ws, col_widths)

    info   = d["1. Informations sur le traitement"][0]
    sec2   = d["2. Données collectées et leur catégories"][0]
    sec3   = d["3. Catégories des données collectées et traitées"]
    sec4   = d["4. Conservation des données"][0]
    sec5   = d["5. Sécurité des traitements et des données"][0]
    sec6   = d["6. Interconnexion et/ou communication des données collectées à des tiers"]
    sec7   = d["7. Transfert des données à l'étranger"][0]
    sec8   = d["8. Consentement des personnes concernées"][0]
    sec9   = d["9. Droits des personnes concernées"]

    r = 1  # current row pointer

    # ── TITLE BLOCK ──────────────────────────────────────────────────────────
    set_row_height(ws, r, 6);  r += 1  # row 1 spacer
    set_row_height(ws, r, 6);  r += 1  # row 2 spacer
    set_row_height(ws, r, 6);  r += 1  # row 3 spacer

    # Row 4: big title
    set_row_height(ws, r, 36)
    ws.merge_cells(f'D{r}:F{r}')
    c = ws.cell(r, 4, 'Accompagnement à la mise en conformité avec la loi 18-07')
    c.font = Font(name="Calibri", bold=True, size=18)
    c.alignment = center()
    ws.cell(r, 7, 'LOGO').font = Font(name="Calibri", bold=True, size=14, color=DARK_RED_FONT)
    r += 1

    # Row 5: subtitle
    set_row_height(ws, r, 30)
    ws.merge_cells(f'D{r}:F{r}')
    c = ws.cell(r, 4, 'Registre des traitements')
    c.font = big_font(22)
    c.alignment = center()
    r += 1

    # Row 6: spacer
    set_row_height(ws, r, 10); r += 1

    # Row 7: label (Déclaration / Autorisation)
    set_row_height(ws, r, 26)
    ws.merge_cells(f'D{r}:F{r}')
    c = ws.cell(r, 4, label)
    c.font = big_font(18)
    c.alignment = center()
    r += 1

    # Rows 8–9: spacers
    set_row_height(ws, r, 8); r += 1
    set_row_height(ws, r, 8); r += 1

    # ── SECTION 1: INFORMATIONS SUR LE TRAITEMENT ────────────────────────────
    set_row_height(ws, r, 30)
    ws.merge_cells(f'A{r}:O{r}')
    style_section_title(ws.cell(r, 1), '1. Informations sur le traitement')
    r += 1

    # Spacer
    r += 1

    # Header row
    set_row_height(ws, r, 28)
    headers_s1 = [
        'Dénomination du traitement','Dénomination du traitement (AR)',
        'Type de traitement','Date de mise en œuvre du traitement',
        'Finalité (but) du traitement ','Finalité (but) du traitement (AR)',
        'Cadre légal du traitement','Cadre légal du traitement (AR)',
        'Catégories des traitements ','','Précisez les autres catégories',
        'Existence de sous traitements',"Existence d'un sous traitant"
    ]
    for ci, h in enumerate(headers_s1, 1):
        style_header_cell(ws.cell(r, ci), h)
    # Merge I:J for categories
    ws.merge_cells(f'I{r}:J{r}')
    r += 1

    # Data row
    set_row_height(ws, r, 22)
    vals_s1 = [
        info.get("Dénomination du traitement",""),
        info.get("Dénomination du traitement (AR)",""),
        info.get("Type de traitement",""),
        info.get("Date de mise en œuvre du traitement",""),
        info.get("Finalité (but) du traitement ",""),
        info.get("Finalité (but) du traitement (AR)",""),
        info.get("Cadre légal du traitement",""),
        info.get("Cadre légal du traitement (AR)",""),
        info.get("Catégories des traitements ",""),"",
        info.get("Précisez les autres catégories",""),
        info.get("Existence de sous traitements",""),
        info.get("Existence d'un sous traitant",""),
    ]
    for ci, v in enumerate(vals_s1, 1):
        style_data_cell(ws.cell(r, ci), v)
    ws.merge_cells(f'I{r}:J{r}')
    r += 1

    # Spacer
    r += 1

    # ── SOUS-TRAITEMENTS ─────────────────────────────────────────────────────
    sts = d["1. Informations sur le traitement"][1:]
    if sts:
        # Sub-traitement header
        set_row_height(ws, r, 28)
        st_headers = [
            'Dénomination du sous traitement','Dénomination du sous traitement (AR)',
            'Type du sous traitement','Objectifs','Sous traité','Observations'
        ]
        for ci, h in enumerate(st_headers, 1):
            style_header_cell(ws.cell(r, ci), h)
        r += 1

        # Group sts by pairs (sous-traitement + sous-traitant)
        i = 0
        alt = False
        while i < len(sts):
            st = sts[i]
            set_row_height(ws, r, 20)
            vals = [
                st.get("Dénomination du sous traitement",""),
                st.get("Dénomination du sous traitement (AR)",""),
                st.get("Type du sous traitement",st.get("Type de traitement","")),
                st.get("Objectifs",""),
                st.get("Sous traité",""),
                st.get("Observations",""),
            ]
            for ci, v in enumerate(vals, 1):
                style_data_cell(ws.cell(r, ci), v, alt)
            r += 1

            # Sous-traitant coords header
            set_row_height(ws, r, 28)
            st_coord_headers = [
                'Type de personne','Nom/Raison sociale','Nom/Raison sociale (AR)',
                'Prénom/Sigle','Prénom/Sigle (AR)','Adresse','Adresse (AR)',
                'Pays','Pays (AR)','Ville','Ville (AR)','N° Tél','N° Fax',
                "Domaine d'activité","Domaine d'activité (AR)",'Site web',
                "Existence d'un contrat signé avec le sous traitant "
            ]
            for ci, h in enumerate(st_coord_headers, 1):
                style_header_cell(ws.cell(r, ci), h)
            r += 1

            set_row_height(ws, r, 20)
            coord_vals = [
                st.get("Type de personne",""), st.get("Nom/Raison sociale",""),
                st.get("Nom/Raison sociale (AR)",""), st.get("Prénom/Sigle",""),
                st.get("Prénom/Sigle (AR)",""), st.get("Adresse",""),
                st.get("Adresse (AR)",""), st.get("Pays",""),
                st.get("Pays (AR)",""), st.get("Ville",""),
                st.get("Ville (AR)",""), st.get("N° Tél",""),
                st.get("N° Fax",""), st.get("Domaine d'activité",""),
                st.get("Domaine d'activité (AR)",""), st.get("Site web",""),
                st.get("Existence d'un contrat signé avec le sous traitant",""),
            ]
            for ci, v in enumerate(coord_vals, 1):
                style_data_cell(ws.cell(r, ci), v, not alt)
            r += 1
            i += 1
            alt = not alt
        r += 1

    # ── SECTION 2: DONNÉES COLLECTÉES ────────────────────────────────────────
    set_row_height(ws, r, 30)
    ws.merge_cells(f'A{r}:O{r}')
    style_section_title(ws.cell(r, 1), '2. Données collectées et leur catégories')
    r += 1; r += 1

    s2_headers = [
        'Type de collecte de données','Mode de collecte',
        'Sécurité existante pour la collecte des données',
        'Catégories des personnes concernées','Autres catégories',
        "Existe-il d'autres sources de données utilisées dans la collecte des données ?",
        'Nom (BD) ou (FM)','Nom structure propriétaire','Moyen de collecte',
        "Cadre légal d'utilisation",'Objectifs '
    ]
    set_row_height(ws, r, 28)
    for ci, h in enumerate(s2_headers, 1):
        style_header_cell(ws.cell(r, ci), h)
    r += 1

    # Build security string
    sec_names = {
        'trac_collect':'Traçabilité','signelect_collect':'Signature électronique',
        'chiffr_collect':'Chiffrement','charte_collect':'Charte de sécurité'
    }
    sec_str = ', '.join(n for k,n in sec_names.items() if sec2.get(k)==1)
    # Build cat persons string
    cat_names = {
        'cat_salaries_collect':'Salariés','cat_usages_collect':'Usagés',
        'cat_patients_collect':'Patients','cat_adherant_collect':'Adhérant',
        'cat_fournisseurs_collect':'Fournisseurs','cat_etudiant_collect':'Étudiants/Élèves',
        'cat_client_collect':'Clients'
    }
    cat_str = ', '.join(n for k,n in cat_names.items() if sec2.get(k)==1)
    if sec2.get('cat_autre_collect')==1 and sec2.get('Autres catégories'):
        cat_str += (', ' if cat_str else '') + sec2['Autres catégories']

    set_row_height(ws, r, 20)
    s2_vals = [
        sec2.get("Type de collecte de données",""),
        sec2.get("Mode de collecte",""),
        sec_str, cat_str,
        sec2.get("Autres catégories",""),
        sec2.get("Existe-il d'autres sources de données utilisées dans la collecte des données ?",""),
        sec2.get("Nom (BD) ou (FM)",""),
        sec2.get("Nom structure propriétaire",""),
        sec2.get("Moyen de collecte",""),
        sec2.get("Cadre légal d'utilisation",""),
        sec2.get("Objectifs ",""),
    ]
    for ci, v in enumerate(s2_vals, 1):
        style_data_cell(ws.cell(r, ci), v)
    r += 1; r += 1

    # ── SECTION 3: CATÉGORIES DES DONNÉES ────────────────────────────────────
    set_row_height(ws, r, 30)
    ws.merge_cells(f'A{r}:O{r}')
    style_section_title(ws.cell(r, 1), '3. Catégories des données collectées et traitées')
    r += 1; r += 1

    s3_headers = [
        'Catégorie de données',"Type d'informations",
        "Autres types d'informations recueillis",'Origine de la donnée',
        'Autres origines de la source',
        'Est elle utilisée pour la finalité du traitement ?',
        'Sources de données','Autres sources',
        'Durée de conservation de la donnée (mois)'
    ]
    set_row_height(ws, r, 28)
    for ci, h in enumerate(s3_headers, 1):
        style_header_cell(ws.cell(r, ci), h)
    r += 1

    for idx, row_data in enumerate(sec3):
        set_row_height(ws, r, 20)
        vals = [
            row_data.get("Catégorie de données",""),
            row_data.get("Type d'informations",""),
            row_data.get("Autres types d'informations recueillis",""),
            row_data.get("Origine de la donnée",""),
            row_data.get("Autres origines de la source",""),
            row_data.get("Est elle utilisée pour la finalité du traitement ?",""),
            row_data.get("Sources de données",""),
            row_data.get("Autres sources",""),
            row_data.get("Durée de conservation de la donnée (mois)",""),
        ]
        for ci, v in enumerate(vals, 1):
            style_data_cell(ws.cell(r, ci), v, idx%2==1)
        r += 1
    r += 1

    # ── SECTION 4: CONSERVATION ───────────────────────────────────────────────
    set_row_height(ws, r, 30)
    ws.merge_cells(f'A{r}:O{r}')
    style_section_title(ws.cell(r, 1), '4. Conservation des données')
    r += 1; r += 1

    s4_headers = [
        'Comment les données sont conservées ?','Nom de la base de données',
        'Lieu de stockage de la base de données','Nom du fichier manuel',
        'Lieu de stockage du fichier'
    ]
    set_row_height(ws, r, 28)
    for ci, h in enumerate(s4_headers, 1):
        style_header_cell(ws.cell(r, ci), h)
    r += 1
    set_row_height(ws, r, 20)
    for ci, v in enumerate([
        sec4.get("Comment les données sont conservées ?",""),
        sec4.get("Nom de la base de données",""),
        sec4.get("Lieu de stockage de la base de données",""),
        sec4.get("Nom du fichier manuel",""),
        sec4.get("Lieu de stockage du fichier",""),
    ], 1):
        style_data_cell(ws.cell(r, ci), v)
    r += 1; r += 1

    # ── SECTION 5: SÉCURITÉ ───────────────────────────────────────────────────
    set_row_height(ws, r, 30)
    ws.merge_cells(f'A{r}:O{r}')
    style_section_title(ws.cell(r, 1), '5. Sécurité des traitements et des données')
    r += 1; r += 1

    s5_headers = [
        'Existe il une charte de sécurité ?',
        'Si oui, la charte de sécurité est-elle lue et signée par le personnel habilité à accéder aux données? ',
        'Sécurité des données, disponibilité de'
    ]
    set_row_height(ws, r, 28)
    for ci, h in enumerate(s5_headers, 1):
        style_header_cell(ws.cell(r, ci), h)
    r += 1
    sec_disp_names = {
        'datacenter_securite':'DATACENTER','backup_securite':'Disaster Recovery',
        'telesurveillance_securite':'Télésurveillance','securite_poste_securite':'Sécurité Postes de Travail',
        'politique_sauv_securite':'Politique de Sauvegarde','chiffrement_securite':'Chiffrement/Déchiffrement',
        'tracabilite_securite':"Traçabilité d'accès",'documentation_securite':'Documentation des procédures',
        'securite_acces_securite':'Sécurité accès physique','mesure_securite':'Mesures sécurité fichier manuel'
    }
    sec_disp = ', '.join(n for k,n in sec_disp_names.items() if sec5.get(k)==1)
    set_row_height(ws, r, 20)
    for ci, v in enumerate([sec5.get("engagement_securite","Oui"), sec5.get("signe_securite","Oui"), sec_disp], 1):
        style_data_cell(ws.cell(r, ci), v)
    r += 1; r += 1

    # ── SECTION 6: INTERCONNEXION ─────────────────────────────────────────────
    set_row_height(ws, r, 30)
    ws.merge_cells(f'A{r}:O{r}')
    style_section_title(ws.cell(r, 1), '6. Interconnexion et/ou communication des données collectées à des tiers')
    r += 1; r += 1

    s6_headers = [
        'Est-ce que vous communiquez des données à caractère personnel à des tiers ou avez des interconnexions ?',
        "Nom de l'organisme destinataire",'Objectifs','Mode de communication','Cadre légal'
    ]
    set_row_height(ws, r, 28)
    for ci, h in enumerate(s6_headers, 1):
        style_header_cell(ws.cell(r, ci), h)
    r += 1
    for idx, entry in enumerate(sec6):
        set_row_height(ws, r, 20)
        for ci, v in enumerate([
            entry.get("Est-ce que vous communiquez des données à caractère personnel à des tiers ou avez des interconnexions ?",""),
            entry.get("Nom de l'organisme destinataire",""),
            entry.get("Objectifs",""),
            entry.get("Mode de communication",""),
            entry.get("Cadre légal",""),
        ], 1):
            style_data_cell(ws.cell(r, ci), v, idx%2==1)
        r += 1
    r += 1

    # ── SECTION 7 & 8: TRANSFERT + CONSENTEMENT ──────────────────────────────
    set_row_height(ws, r, 30)
    ws.merge_cells(f'A{r}:G{r}')
    style_section_title(ws.cell(r, 1), "7. Transfert des données à l'étranger")
    ws.merge_cells(f'H{r}:O{r}')
    style_section_title(ws.cell(r, 8), "8. Consentement des personnes concernées")
    r += 1; r += 1

    s78_headers = [
        'Les données traitées sont-elles transférées vers un pays étranger?',
        'Consentement des personnes concernées ','Méthode de consentement',
        'Méthode de consentement (AR)',"Précisez pourquoi y'a pas de consentement des personnes"
    ]
    set_row_height(ws, r, 28)
    for ci, h in enumerate(s78_headers, 1):
        style_header_cell(ws.cell(r, ci), h)
    r += 1
    set_row_height(ws, r, 20)
    for ci, v in enumerate([
        sec7.get("Les données traitées sont-elles transférées vers un pays étranger?","Non"),
        sec8.get("Consentement des personnes concernées ",""),
        sec8.get("Méthode de consentement",""),
        sec8.get("Méthode de consentement (AR)",""),
        sec8.get("Précisez pourquoi y'a pas de consentement des personnes",""),
    ], 1):
        style_data_cell(ws.cell(r, ci), v)
    r += 1; r += 1; r += 1

    # ── SECTION 9: DROITS DES PERSONNES ──────────────────────────────────────
    set_row_height(ws, r, 30)
    ws.merge_cells(f'A{r}:O{r}')
    style_section_title(ws.cell(r, 1), '9. Droits des personnes concernées ')
    r += 1; r += 1

    s9_headers = [
        'Les personnes concernées sont elles informées sur le contenu du traitement de ',
        'Comment','Comment (AR)',
        "le nom du service auprès duquel la personne concernée pourra exercer le droit à l'information/l'accès/  la rectification / l'opposition",
        "le nom du service auprès duquel la personne concernée pourra exercer le droit à l'information/l'accès/  la rectification / l'opposition (AR)",
        'Adresse','Adresse (AR)','Mobile','Fax','e-Mail',
        "Quelles sont les mesures prises pour faciliter l'exercice du droit d'information / d'accès / de rectification / d'opposition",
        "Quelles sont les mesures prises pour faciliter l'exercice du droit d'information / d'accès / de rectification / d'opposition (AR)"
    ]
    set_row_height(ws, r, 28)
    for ci, h in enumerate(s9_headers, 1):
        style_header_cell(ws.cell(r, ci), h)
    r += 1

    articles = [
        "l'article 32 de la loi 18-07","l'article 34 de la loi 18-07",
        "l'article 35 de la loi 18-07","l'article 36 de la loi 18-07"
    ]
    droits_row = sec9[0] if sec9 else {}
    for idx, art in enumerate(articles):
        set_row_height(ws, r, 20)
        for ci, v in enumerate([
            art,
            droits_row.get("Comment",""),
            droits_row.get("Comment (AR)",""),
            droits_row.get("le nom du service auprès duquel la personne concernée pourra exercer le droit à l'information/l'accès/  la rectification / l'opposition",""),
            droits_row.get("le nom du service auprès duquel la personne concernée pourra exercer le droit à l'information/l'accès/  la rectification / l'opposition (AR)",""),
            droits_row.get("Adresse",""),
            droits_row.get("Adresse (AR)",""),
            droits_row.get("Mobile",""),
            droits_row.get("Fax",""),
            droits_row.get("e-Mail",""),
            droits_row.get("Quelles sont les mesures prises pour faciliter l'exercice du droit d'information / d'accès / de rectification / d'opposition",""),
            droits_row.get("Quelles sont les mesures prises pour faciliter l'exercice du droit d'information / d'accès / de rectification / d'opposition (AR)",""),
        ], 1):
            style_data_cell(ws.cell(r, ci), v, idx%2==1)
        r += 1

    # Freeze panes (freeze top rows)
    ws.freeze_panes = 'A11'
    return ws


def build_choix_sheet(wb):
    ws = wb.create_sheet(title='choix')
    choix_data = [
        ['Type de traitement/Type de sous traitement','Catégories des traitements',
         'Sécurité existante pour la collecte des données','Catégories des personnes concernées',
         'Sécurité des données, disponibilité de','Conservation de la donnée',
         'mode de com','','personnelles','professionnelles','financières','sensibles'],
        ['Automatique','Gestion du Personnel','Traçabilité','Salaries','DATACENTER','Informatique','Papier','','NIN','Fonction','Revenu','Origine raciale ou ethnique'],
        ['Manuel','Comptabilité','Signature électronique','Usagés','Disaster Recovery','Manuel','Support externe','','Nom et prénom','Employeur','Compte bancaire','Opinions politiques'],
        ['Automatique, Manuel','Service public','Chiffrement','Patients','Système de télésurveillance','','Connexion','','Date de naissance','Curriculum vitae','Autre à préciser','Convictions religieuses ou philosophiques'],
        ['','Gestion des Patients','Charte de sécurité','Adhérant','Sécurité des postes de travail','','','','Lieu de naissance','Autre à préciser','','Appartenance syndicale'],
        ['','Gestion des Étudiants/Élèves','','Fournisseurs','Politique de sauvegarde des données','','à compléter','','Situation de famille','','','Données de santé'],
        ['','Recherche scientifique','','Étudiants/Élèves','Chiffrement et Déchiffrement des données','','','','Photos','','','Données génétiques'],
        ['','Sécurité et contrôle d\'accès','','Clients (actuels ou potentiels)','Traçabilité d\'accès  aux données','','','','Données biométriques'],
        ['','Gestion des Fournisseurs','','Autres','Documentation des procédures de sécurité','','','','Adresse du domicile'],
        ['','Gestion des Clients','','','Sécurité d\'accès physique aux locaux','','','','Adresse mail'],
        ['','Autres','','','Mesures de sécurité du fichier manuel','','','','N° téléphone'],
        ['','','','','','','','','N° de la pièce d\'identité/Permis/Passeport'],
        ['','','','','','','','','Autre à préciser'],
    ]
    for ri, row in enumerate(choix_data, 1):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            if ri == 1:
                c.font = hdr_font(9)
                c.fill = green_fill()
                c.alignment = center()
                c.border = all_borders()
            else:
                c.font = data_font(9)
                c.alignment = center()
                c.border = all_borders()
    for col in ['A','B','C','D','E','F','G','H','I','J','K','L']:
        ws.column_dimensions[col].width = 30
    return ws


def generate(json_path, out_path):
    with open(json_path, encoding='utf-8') as f:
        d = json.load(f)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_sheet(wb, 'Autorisation', d, 'Autorisation')
    build_sheet(wb, 'declaration', d, 'Déclaration')
    build_choix_sheet(wb)

    wb.save(out_path)
    print(f"OK:{out_path}")


if __name__ == '__main__':
    generate(sys.argv[1], sys.argv[2])
