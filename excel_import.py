"""
Excel → Traitement data extractor
Based on the real xlsm structure:
- One sheet per traitement (skip: PROSERVE/company name, choix, Autorisation, declaration, Declaration_)
- Each sheet has sections as row headers, data in subsequent rows
"""
import re
import openpyxl
import json

SKIP_SHEETS = {'choix', 'autorisation', 'declaration', 'declaration_', 'choix'}

CATEGORY_MAP = {
    "personnelles":    "Données à caractère personnel",
    "professionnelles":"Données professionnelles",
    "financières":     "Données financières",
    "sensibles":       "Données  sensibles",
}

SECURITY_MAP = {
    "datacenter":                           "datacenter_securite",
    "disaster recovery":                    "backup_securite",
    "système de télésurveillance":          "telesurveillance_securite",
    "sécurité des postes de travail":       "securite_poste_securite",
    "politique de sauvegarde des données":  "politique_sauv_securite",
    "traçabilité d'accès  aux données":     "tracabilite_securite",
    "documentation des procédures de sécurité": "documentation_securite",
    "sécurité d'accès physique aux locaux": "securite_acces_securite",
    "mesures de sécurité du fichier manuel":"mesure_securite",
    "chiffrement et déchiffrement des données": "chiffrement_securite",
}

COLLECT_SEC_MAP = {
    "traçabilité":           "trac_collect",
    "signature électronique":"signelect_collect",
    "chiffrement":           "chiffr_collect",
    "charte de sécurité":    "charte_collect",
}

COLLECT_CAT_MAP = {
    "salaries":                    "cat_salaries_collect",
    "usagés":                      "cat_usages_collect",
    "patients":                    "cat_patients_collect",
    "adhérant":                    "cat_adherant_collect",
    "fournisseurs":                "cat_fournisseurs_collect",
    "étudiants/élèves":            "cat_etudiant_collect",
    "clients (actuels ou potentiels)": "cat_client_collect",
    "autres":                      "cat_autre_collect",
}

def v(cell_val):
    """Clean cell value."""
    if cell_val is None:
        return ""
    s = str(cell_val).strip()
    return "" if s.lower() in ("n/a", "none", "à compléter", "a compléter") else s

def is_section_header(row_vals, keyword):
    """Check if first non-empty cell starts a section."""
    for c in row_vals:
        if c and keyword.lower() in str(c).lower():
            return True
    return False

def row_to_dict(header_row, data_row):
    """Map header cells to data cells."""
    d = {}
    for h, val in zip(header_row, data_row):
        if h and str(h).strip():
            d[str(h).strip()] = v(val)
    return d

def parse_sheet(ws):
    """
    Parse one sheet into a traitement data dict.
    Returns (name, data_dict, log_list)
    """
    log = []

    # Read all rows as lists
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([v(c) for c in row])

    # Find the traitement name (first non-empty cell in section 1 data row)
    name = ws.title.strip()

    # ── State machine: find each section ─────────────────────────────────────
    sec1_info = {}
    sous_traitements = []
    current_sous = None
    current_sous_traitants = []
    sec2 = {}
    sec3_rows = []
    sec4 = {}
    sec5 = {}
    sec6_rows = []
    sec7 = {}
    sec8 = {}
    sec9_rows = []

    i = 0
    while i < len(rows):
        row = rows[i]
        first = row[0] if row else ""

        # ── Section 1 header ─────────────────────────────────────────────────
        if "1. Informations sur le traitement" in first:
            i += 1
            # Skip blank rows to find column header row
            while i < len(rows) and not any(rows[i]):
                i += 1
            if i < len(rows):
                header = rows[i]
                i += 1
                # Skip blank rows to find data row
                while i < len(rows) and not any(rows[i]):
                    i += 1
                if i < len(rows):
                    data = rows[i]
                    sec1_info = row_to_dict(header, data)
                    if sec1_info.get("Dénomination du traitement"):
                        name = sec1_info["Dénomination du traitement"]

                    # Read subsequent sous-traitement blocks
                    i += 1
                    while i < len(rows):
                        r = rows[i]
                        if r[0] and ("2." in r[0] or "Données collectées" in r[0]):
                            break
                        # Sous-traitement header row
                        if r[0] == "Dénomination du sous traitement":
                            i += 1
                            if i < len(rows):
                                st_data = row_to_dict(r, rows[i])
                                current_sous = {
                                    "Dénomination du sous traitement": st_data.get("Dénomination du sous traitement",""),
                                    "Dénomination du sous traitement (AR)": st_data.get("Dénomination du sous traitement (AR)",""),
                                    "Type du sous traitement": st_data.get("Type du sous traitement",""),
                                    "Objectifs": st_data.get("Objectifs",""),
                                    "Sous traité": st_data.get("Sous traité","Non"),
                                    "Observations": st_data.get("Observations",""),
                                    "sous_traitants": []
                                }
                                sous_traitements.append(current_sous)
                        # Sous-traitant header row
                        elif r[0] == "Type de personne":
                            i += 1
                            if i < len(rows) and rows[i][0]:
                                st_data = row_to_dict(r, rows[i])
                                if current_sous:
                                    current_sous["sous_traitants"].append(st_data)
                        i += 1
                    continue

        # ── Section 2 ─────────────────────────────────────────────────────────
        elif "2. Données collectées" in first:
            i += 1
            while i < len(rows) and not any(rows[i]): i += 1
            if i < len(rows):
                header = rows[i]
                i += 1
                while i < len(rows) and not any(rows[i]): i += 1
                if i < len(rows):
                    data = rows[i]
                    raw = row_to_dict(header, data)

                    # Type de collecte
                    tc = raw.get("Type de collecte de données","").replace("_", " ")
                    sec2["Type de collecte de données"] = tc

                    # Mode de collecte
                    sec2["Mode de collecte"] = raw.get("Mode de collecte","")

                    # Security of collection
                    sec_raw = raw.get("Sécurité existante pour la collecte des données","")
                    selected = {s.strip().lower() for s in sec_raw.split(",") if s.strip()}
                    for label, key in COLLECT_SEC_MAP.items():
                        sec2[key] = "1" if label.lower() in selected else "0"

                    # Categories of persons
                    cat_raw = raw.get("Catégories des personnes concernées","")
                    selected_cat = {s.strip().lower() for s in cat_raw.split(",") if s.strip()}
                    for label, key in COLLECT_CAT_MAP.items():
                        sec2[key] = "1" if label.lower() in selected_cat else "0"

                    sec2["Autres catégories"] = raw.get("Autres catégories","")
                    sec2["Existe-il d'autres sources de données utilisées dans la collecte des données ?"] = raw.get("Existe-il d'autres sources de données utilisées dans la collecte des données ?","Non")
                    sec2["Nom (BD) ou (FM)"] = raw.get("Nom (BD) ou (FM)","")
                    sec2["Nom structure propriétaire"] = raw.get("Nom structure propriétaire","")

        # ── Section 3 ─────────────────────────────────────────────────────────
        elif "3. Catégories des données" in first:
            i += 1
            while i < len(rows) and not any(rows[i]): i += 1
            if i < len(rows):
                header = rows[i]
                i += 1
                while i < len(rows):
                    r = rows[i]
                    if r[0] and ("4." in r[0] or "Conservation" in r[0]):
                        break
                    if not r[0]:
                        i += 1
                        continue
                    raw = row_to_dict(header, r)
                    cat_val = CATEGORY_MAP.get(raw.get("Catégorie de données","").lower().strip(),
                                               raw.get("Catégorie de données",""))

                    # Split comma-separated types into individual rows
                    types_raw = raw.get("Type d'informations","")
                    types = [t.strip() for t in types_raw.split(",") if t.strip()]
                    autres_raw = raw.get("Autres types d'informations recueillis","")
                    autres = [a.strip() for a in autres_raw.split(",")
                              if a.strip() and a.strip().lower() not in ("n/a","")]

                    for t in (types or [""]):
                        entry = {
                            "Catégorie de données": cat_val,
                            "Type d'informations": t,
                            "Autres types d'informations recueillis": autres_raw,
                            "Origine de la donnée": raw.get("Origine de la donnée","De la personne concernée"),
                            "Autres origines de la source": raw.get("Autres origines de la source",""),
                            "Est elle utilisée pour la finalité du traitement ?": raw.get("Est elle utilisée pour la finalité du traitement ?","Oui"),
                            "Sources de données": raw.get("Sources de données",""),
                            "Autres sources": raw.get("Autres sources",""),
                            "Durée de conservation de la donnée (mois)": re.sub(r"[^\d]","", raw.get("Durée de conservation de la donnée (mois)","0")) or "0",
                        }
                        # Fix "Origine de la donnée"
                        if entry["Origine de la donnée"] in ("Personne concernée","Pers. concernée"):
                            entry["Origine de la donnée"] = "De la personne concernée"
                            log.append(f"Section 3 ({cat_val}): Origine corrigée → 'De la personne concernée'")

                        sec3_rows.append(entry)
                    i += 1
                continue

        # ── Section 4 / 5 (Conservation) ─────────────────────────────────────
        elif ("4. Conservation" in first or "5. Conservation" in first):
            i += 1
            if i < len(rows):
                header = rows[i]
                i += 1
                if i < len(rows):
                    data = rows[i]
                    raw = row_to_dict(header, data)
                    sec4["Comment les données sont conservées ?"] = raw.get("Comment les données sont conservées ?","")
                    sec4["Nom de la base de données"] = raw.get("Nom de la base de données","")
                    sec4["Lieu de stockage de la base de données"] = raw.get("Lieu de stockage de la base de données","")
                    sec4["Nom du fichier manuel"] = raw.get("Nom du fichier manuel","")
                    sec4["Lieu de stockage du fichier"] = raw.get("Lieu de stockage du fichier","")

        # ── Section 5 / 6 (Sécurité) ─────────────────────────────────────────
        elif ("5. Sécurité" in first or "6. Sécurité" in first):
            i += 1
            if i < len(rows):
                header = rows[i]
                i += 1
                if i < len(rows):
                    data = rows[i]
                    raw = row_to_dict(header, data)
                    sec5["engagement_securite"] = raw.get("Existe il une charte de sécurité ?","Oui")

                    sec_raw = raw.get("Sécurité des données, disponibilité de","")
                    selected = {s.strip().lower() for s in sec_raw.split(",") if s.strip()}
                    for label, key in SECURITY_MAP.items():
                        sec5[key] = "1" if label.lower() in selected else "0"

        # ── Section 6 / 7 (Interconnexion) ───────────────────────────────────
        elif ("6. Interconnexion" in first or "7. Interconnexion" in first):
            i += 1
            if i < len(rows):
                header = rows[i]
                i += 1
                while i < len(rows):
                    r = rows[i]
                    if r[0] and ("7." in r[0] or "8." in r[0] or "Transfert" in r[0] or "Consentement" in r[0]):
                        break
                    if r[0] and r[0].lower() not in ("oui","non",""):
                        i += 1
                        continue
                    if r[0]:
                        raw = row_to_dict(header, r)
                        sec6_rows.append({
                            "Est-ce que vous communiquez des données à des tiers ?": raw.get("Est-ce que vous communiquez des données à des tiers ?","Non"),
                            "Nom de l'organisme destinataire": raw.get("Nom de l'organisme destinataire",""),
                            "Objectifs": raw.get("Objectifs",""),
                            "Mode de communication": raw.get("Mode de communication",""),
                            "Cadre légal": raw.get("Cadre légal",""),
                        })
                    i += 1
                continue

        # ── Section 7+8 / 8+9 (Transfert + Consentement) ─────────────────────
        elif ("7. Transfert" in first or "8. Transfert" in first):
            i += 1
            if i < len(rows):
                header = rows[i]
                i += 1
                if i < len(rows):
                    data = rows[i]
                    raw = row_to_dict(header, data)
                    sec7["Les données traitées sont-elles transférées vers un pays étranger?"] = raw.get("Les données traitées sont-elles transférées vers un pays étranger?","Non")
                    # Consentement (same row, cols 4+)
                    sec8["Consentement des personnes concernées "] = raw.get("Consentement des personnes concernées ","Oui")
                    sec8["Méthode de consentement"] = raw.get("Méthode de consentement","")
                    sec8["Méthode de consentement (AR)"] = raw.get("Méthode de consentement (AR)","")
                    sec8["Précisez pourquoi y'a pas de consentement"] = raw.get("Précisez pourquoi y'a pas de consentement","")

        # ── Section 9 (Droits) ────────────────────────────────────────────────
        elif ("9. Droits" in first):
            i += 1
            # Skip blank rows to find header
            while i < len(rows) and not any(rows[i]): i += 1
            if i < len(rows):
                i += 1  # skip header row (we'll read by position)
                while i < len(rows):
                    r = rows[i]
                    if r[0] and "10." in str(r[0]):
                        break
                    if not any(r):
                        i += 1
                        continue
                    row_vals = [v(x) for x in r]
                    if any(row_vals):
                        entry = {
                            "article": row_vals[0] if len(row_vals) > 0 else "",
                            "Oui/Non": row_vals[1] if len(row_vals) > 1 else "Oui",
                            "Comment": row_vals[2] if len(row_vals) > 2 else "",
                            "Comment (AR)": row_vals[3] if len(row_vals) > 3 else "",
                            "le nom du service auprès duquel la personne concernée pourra exercer le droit à l'information/l'accès/  la rectification / l'opposition": row_vals[4] if len(row_vals) > 4 else "",
                            "le nom du service auprès duquel la personne concernée pourra exercer le droit à l'information/l'accès/  la rectification / l'opposition (AR)": row_vals[5] if len(row_vals) > 5 else "",
                            "Adresse": row_vals[6] if len(row_vals) > 6 else "",
                            "Adresse (AR)": row_vals[7] if len(row_vals) > 7 else "",
                            "Mobile": row_vals[8] if len(row_vals) > 8 else "",
                            "Fax": row_vals[9] if len(row_vals) > 9 else "",
                            "e-Mail": row_vals[10] if len(row_vals) > 10 else "",
                            "Quelles sont les mesures prises pour faciliter l'exercice du droit d'information / d'accès / de rectification / d'opposition": row_vals[11] if len(row_vals) > 11 else "",
                            "Quelles sont les mesures prises pour faciliter l'exercice du droit d'information / d'accès / de rectification / d'opposition (AR)": row_vals[12] if len(row_vals) > 12 else "",
                        }
                        sec9_rows.append(entry)
                    i += 1
                continue

        i += 1

    # ── Build sous-traitements for section 1 ─────────────────────────────────
    sec1_sous = []
    for st in sous_traitements:
        traitants = st.pop("sous_traitants", [])
        sec1_sous.append(st)
        for tr in traitants:
            sec1_sous.append(tr)

    # ── Assemble final data dict ──────────────────────────────────────────────
    data = {
        "1. Informations sur le traitement": [sec1_info] + sec1_sous,
        "2. Données collectées et leur catégories": [sec2] if sec2 else [],
        "3. Catégories des données collectées et traitées": sec3_rows,
        "4. Conservation des données": [sec4] if sec4 else [],
        "5. Sécurité des traitements et des données": [sec5] if sec5 else [],
        "6. Interconnexion et/ou communication des données collectées à des tiers": sec6_rows,
        "7. Transfert des données à l'étranger": [sec7] if sec7 else [],
        "8. Consentement des personnes concernées": [sec8] if sec8 else [],
        "9. Droits des personnes concernées": sec9_rows,
    }

    if not log:
        log.append("✅ Aucune correction — données importées telles quelles")
    else:
        log.insert(0, f"⚠ {len(log)} correction(s) appliquée(s)")

    return name, data, log


def extract_from_excel(file_path):
    """
    Returns list of (name, data_dict, log_list) — one per traitement sheet.
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True, keep_vba=True)
    except Exception:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        row_count = sum(1 for row in ws.iter_rows(min_row=1, max_row=5, values_only=True)
                       if any(cell for cell in row))
        if row_count < 1:
            continue
        try:
            name, data, log = parse_sheet(ws)
            # Skip blank template sheets
            info = (data.get('1. Informations sur le traitement') or [{}])[0]
            if not any([info.get('Dénomination du traitement'),
                        info.get('Finalité (but) du traitement'),
                        info.get('Finalité (but) du traitement '),
                        data.get('3. Catégories des données collectées et traitées')]):
                continue
            results.append((name, data, log))
        except Exception as e:
            results.append((sheet_name,
                {'1. Informations sur le traitement': [{'Dénomination du traitement': sheet_name}]},
                [f'❌ Erreur lecture: {str(e)[:200]}']))

    wb.close()
    return results
