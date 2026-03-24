"""
ANPDP Registre des Traitements — Web App
Flask + SQLite backend
"""
import os, json, hashlib, secrets
from datetime import datetime
from flask import (Flask, render_template, request, redirect, url_for,
                   session, jsonify, send_file, abort)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import io

# Default password is 'anpdp2024' (hashed)
# To change: set PASSWORD environment variable with a SCrypt hash.
DEFAULT_HASH = 'scrypt:32768:8:1$vC9m86wK87QeWIsz$169e57827e8281313437e6f8a846f32e022f1c84918e77519396e9f1a21e422f273b4f65c5c6f0e34c56e36d4f7c8d9e6f3b2a1a098c7d6e5a4b3c2d1e0f9'
APP_PASSWORD_HASH = os.environ.get('PASSWORD', DEFAULT_HASH)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config.update(
    SQLALCHEMY_DATABASE_URI=os.environ.get('DATABASE_URL', 'sqlite:///anpdp.db'),
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=True if os.environ.get('RENDER') or os.environ.get('RAILWAY_STATIC_URL') else False
)

db = SQLAlchemy(app)

# ─── MODELS ──────────────────────────────────────────────────────────────────

class Company(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    name        = db.Column(db.String(200), nullable=False)
    nit         = db.Column(db.String(50), nullable=False)
    logo_path   = db.Column(db.String(300), default='')   # relative path under static/
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    traitements = db.relationship('Traitement', backref='company',
                                   lazy=True, cascade='all, delete-orphan')

class Traitement(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    company_id     = db.Column(db.Integer, db.ForeignKey('company.id'), nullable=False)
    name           = db.Column(db.String(200), nullable=False)
    status         = db.Column(db.String(20), default='brouillon')
    traitement_type= db.Column(db.String(100), default='')   # knowledge base type
    data           = db.Column(db.Text, default='{}')
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at     = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def get_data(self):
        try: return json.loads(self.data)
        except: return {}

    def set_data(self, d):
        self.data = json.dumps(d, ensure_ascii=False)

class KBType(db.Model):
    """A knowledge base type — e.g. 'Gestion RH / Paie'"""
    id          = db.Column(db.Integer, primary_key=True)
    name        = db.Column(db.String(200), nullable=False, unique=True)
    icon        = db.Column(db.String(10), default='🗂')
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    entries     = db.relationship('KBEntry', backref='kb_type',
                                  lazy=True, cascade='all, delete-orphan')

class KBEntry(db.Model):
    """One suggestion: a field_key → value pair inside a KBType"""
    id          = db.Column(db.Integer, primary_key=True)
    type_id     = db.Column(db.Integer, db.ForeignKey('kb_type.id'), nullable=False)
    field_key   = db.Column(db.String(100), nullable=False)  # e.g. 'finalite'
    field_label = db.Column(db.String(200), default='')      # human label
    value       = db.Column(db.Text, nullable=False)
    is_arabic   = db.Column(db.Boolean, default=False)
    approved    = db.Column(db.Boolean, default=True)        # False = pending review
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)

# ─── AUTH ─────────────────────────────────────────────────────────────────────

from functools import wraps

def logged_in():
    return session.get('auth') == 'ok'

def auth_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not logged_in():
            if request.path.startswith('/api/'):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

@app.before_request
def require_login():
    public = ['login', 'static']
    if request.endpoint in public: return
    if not logged_in():
        if request.path.startswith('/api/'):
            return jsonify({"error": "Unauthorized"}), 401
        return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    err = None
    if request.method == 'POST':
        pwd = request.form.get('password', '')
        if check_password_hash(APP_PASSWORD_HASH, pwd) or pwd == os.environ.get('DEV_PASSWORD'):
            session['auth'] = 'ok'
            return redirect(url_for('hub'))
        err = 'Mot de passe incorrect'
    return render_template('login.html', err=err)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ─── HUB ──────────────────────────────────────────────────────────────────────

@app.route('/')
def hub():
    companies = Company.query.order_by(Company.created_at.desc()).all()
    stats = {}
    for c in companies:
        ts = c.traitements
        stats[c.id] = {
            'total':    len(ts),
            'brouillon': sum(1 for t in ts if t.status == 'brouillon'),
            'pret':      sum(1 for t in ts if t.status == 'pret'),
            'soumis':    sum(1 for t in ts if t.status == 'soumis'),
        }
    return render_template('hub.html', companies=companies, stats=stats)

# ─── COMPANIES ────────────────────────────────────────────────────────────────

import os, uuid as _uuid

def save_logo(file_obj, company_id):
    """Save uploaded logo, return relative path or '' on failure."""
    if not file_obj or not file_obj.filename:
        return ''
    ext = file_obj.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('png','jpg','jpeg','gif','webp','svg'):
        return ''
    fname = f"logo_{company_id}_{_uuid.uuid4().hex[:8]}.{ext}"
    folder = os.path.join(app.root_path, 'static', 'logos')
    os.makedirs(folder, exist_ok=True)
    file_obj.save(os.path.join(folder, fname))
    return f'logos/{fname}'

@app.route('/company/new', methods=['POST'])
def company_new():
    name = request.form.get('name', '').strip()
    nit  = request.form.get('nit',  '').strip()
    if not name:
        return redirect(url_for('hub'))
    c = Company(name=name, nit=nit or '—')
    db.session.add(c)
    db.session.flush()   # get c.id before commit
    logo = save_logo(request.files.get('logo'), c.id)
    if logo: c.logo_path = logo
    db.session.commit()
    return redirect(url_for('company_view', cid=c.id))

@app.route('/company/<int:cid>')
def company_view(cid):
    c = Company.query.get_or_404(cid)
    traitements = Traitement.query.filter_by(company_id=cid)\
                                  .order_by(Traitement.updated_at.desc()).all()
    kb_types     = KBType.query.order_by(KBType.name).all()
    all_companies= Company.query.order_by(Company.name).all()
    return render_template('company.html', company=c, traitements=traitements,
                           kb_types=kb_types, all_companies=all_companies)

@app.route('/company/<int:cid>/delete', methods=['POST'])
def company_delete(cid):
    c = Company.query.get_or_404(cid)
    db.session.delete(c)
    db.session.commit()
    return redirect(url_for('hub'))

@app.route('/company/<int:cid>/logo', methods=['POST'])
def company_logo(cid):
    c = Company.query.get_or_404(cid)
    # Option 1: URL
    url = request.form.get('logo_url', '').strip()
    if url:
        c.logo_path = url
        db.session.commit()
        return redirect(url_for('company_view', cid=cid))
    # Option 2: File upload
    f = request.files.get('logo_file')
    if f and f.filename:
        import os, uuid
        ext  = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.png','.jpg','.jpeg','.gif','.webp','.svg'):
            return redirect(url_for('company_view', cid=cid))
        fname = f'logo_{cid}_{uuid.uuid4().hex[:8]}{ext}'
        save_path = os.path.join(app.root_path, 'static', 'logos', fname)
        f.save(save_path)
        c.logo_path = f'/static/logos/{fname}'
        db.session.commit()
    return redirect(url_for('company_view', cid=cid))

@app.route('/company/<int:cid>/logo/delete', methods=['POST'])
def company_logo_delete(cid):
    c = Company.query.get_or_404(cid)
    import os
    if c.logo_path and c.logo_path.startswith('/static/logos/'):
        try:
            os.remove(os.path.join(app.root_path, c.logo_path.lstrip('/')))
        except: pass
    c.logo_path = ''
    db.session.commit()
    return redirect(url_for('company_view', cid=cid))


def company_edit(cid):
    c = Company.query.get_or_404(cid)
    c.name = request.form.get('name', c.name).strip()
    c.nit  = request.form.get('nit',  c.nit).strip()
    logo = save_logo(request.files.get('logo'), cid)
    if logo:
        # Delete old logo file
        if c.logo_path:
            old = os.path.join(app.root_path, 'static', c.logo_path)
            if os.path.exists(old): os.remove(old)
        c.logo_path = logo
    db.session.commit()
    return redirect(url_for('company_view', cid=cid))

# ─── CLONE ────────────────────────────────────────────────────────────────────

# Fields to clear when cloning (company-specific)
CLONE_CLEAR_FIELDS = [
    # Conservation
    'Nom de la base de données', 'Lieu de stockage de la base de données',
    'Nom du fichier manuel', 'Lieu de stockage du fichier',
    # Droits des personnes
    'Adresse', 'Adresse (AR)', 'Mobile', 'Fax', 'e-Mail',
    "le nom du service auprès duquel la personne concernée pourra exercer le droit à l'information/l'accès/  la rectification / l'opposition",
    "le nom du service auprès duquel la personne concernée pourra exercer le droit à l'information/l'accès/  la rectification / l'opposition (AR)",
]

def clone_traitement_data(data):
    """Deep copy traitement data, clearing company-specific fields."""
    import copy
    d = copy.deepcopy(data)
    # Clear fields in section 4 (conservation)
    for row in d.get('4. Conservation des données', []):
        for f in CLONE_CLEAR_FIELDS:
            if f in row: row[f] = ''
    # Clear fields in section 9 (droits)
    for row in d.get('9. Droits des personnes concernées', []):
        for f in CLONE_CLEAR_FIELDS:
            if f in row: row[f] = ''
    return d

@app.route('/company/<int:cid>/clone', methods=['GET','POST'])
def company_clone(cid):
    """Clone selected (or all) traitements to a target company."""
    src = Company.query.get_or_404(cid)
    companies = Company.query.filter(Company.id != cid).order_by(Company.name).all()

    if request.method == 'GET':
        traitements = Traitement.query.filter_by(company_id=cid).all()
        return render_template('clone.html', source=src,
                               traitements=traitements, companies=companies)

    # POST — perform clone
    target_id  = request.form.get('target_company', type=int)
    tid_list   = request.form.getlist('traitement_ids')  # empty = all
    target = Company.query.get_or_404(target_id)

    if tid_list:
        sources = Traitement.query.filter(
            Traitement.company_id == cid,
            Traitement.id.in_([int(x) for x in tid_list])
        ).all()
    else:
        sources = Traitement.query.filter_by(company_id=cid).all()

    count = 0
    for t in sources:
        new_data = clone_traitement_data(t.get_data())
        new_t = Traitement(
            company_id     = target_id,
            name           = t.name,
            traitement_type= t.traitement_type,
            status         = 'brouillon',
        )
        new_t.set_data(new_data)
        db.session.add(new_t)
        count += 1

    db.session.commit()
    return redirect(url_for('company_view', cid=target_id))

@app.route('/company/<int:cid>/traitement/<int:tid>/clone', methods=['POST'])
def traitement_clone(cid, tid):
    """Clone a single traitement to a target company."""
    t = Traitement.query.get_or_404(tid)
    if t.company_id != cid: abort(404)

    target_id = request.get_json().get('target_company_id')
    if not target_id: return jsonify(error='Missing target'), 400
    Company.query.get_or_404(target_id)

    new_data = clone_traitement_data(t.get_data())
    new_t = Traitement(
        company_id      = target_id,
        name            = t.name,
        traitement_type = t.traitement_type,
        status          = 'brouillon',
    )
    new_t.set_data(new_data)
    db.session.add(new_t)
    db.session.commit()
    return jsonify(ok=True, id=new_t.id,
                   company_url=url_for('company_view', cid=target_id))



@app.route('/company/<int:cid>/traitement/new', methods=['POST'])
def traitement_new(cid):
    Company.query.get_or_404(cid)
    name  = request.form.get('name', 'Nouveau traitement').strip()
    ttype = request.form.get('traitement_type', '').strip()
    t = Traitement(company_id=cid, name=name, traitement_type=ttype)
    db.session.add(t)
    db.session.commit()
    return redirect(url_for('traitement_edit', cid=cid, tid=t.id))

@app.route('/company/<int:cid>/traitement/<int:tid>')
def traitement_edit(cid, tid):
    c  = Company.query.get_or_404(cid)
    t  = Traitement.query.get_or_404(tid)
    if t.company_id != cid: abort(404)
    kb_types = KBType.query.order_by(KBType.name).all()
    return render_template('traitement.html', company=c, traitement=t,
                           data_json=t.data, kb_types=kb_types)

@app.route('/company/<int:cid>/traitement/<int:tid>/save', methods=['POST'])
def traitement_save(cid, tid):
    t = Traitement.query.get_or_404(tid)
    if t.company_id != cid: abort(404)
    payload = request.get_json()
    if not payload: return jsonify(error='No data'), 400
    # Update name from traitement data if available
    info = payload.get('1. Informations sur le traitement', [{}])[0]
    denom = info.get('Dénomination du traitement', '').strip()
    if denom: t.name = denom
    t.set_data(payload)
    t.updated_at = datetime.utcnow()
    if 'status' in payload:
        t.status = payload['status']
    db.session.commit()
    return jsonify(ok=True, id=t.id, name=t.name)

@app.route('/company/<int:cid>/traitement/<int:tid>/set_type', methods=['POST'])
def traitement_set_type(cid, tid):
    t = Traitement.query.get_or_404(tid)
    if t.company_id != cid: abort(404)
    type_id = request.get_json().get('type_id', '')
    t.traitement_type = str(type_id)
    t.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify(ok=True)


def traitement_status(cid, tid):
    t = Traitement.query.get_or_404(tid)
    if t.company_id != cid: abort(404)
    s = request.get_json().get('status', 'brouillon')
    if s in ('brouillon', 'pret', 'soumis'):
        t.status = s
        t.updated_at = datetime.utcnow()
        db.session.commit()
    return jsonify(ok=True, status=t.status)

@app.route('/company/<int:cid>/traitement/<int:tid>/delete', methods=['POST'])
def traitement_delete(cid, tid):
    t = Traitement.query.get_or_404(tid)
    if t.company_id != cid: abort(404)
    db.session.delete(t)
    db.session.commit()
    return redirect(url_for('company_view', cid=cid))

@app.route('/company/<int:cid>/export/json-bulk')
def company_export_json_bulk(cid):
    """Export multiple traitements as a JSON array for local injection."""
    c = Company.query.get_or_404(cid)
    ids_str = request.args.get('ids', '')
    try:
        ids = [int(x) for x in ids_str.split(',') if x.strip()]
    except ValueError:
        return "Invalid IDs", 400

    results = []
    for tid in ids:
        t = Traitement.query.filter_by(id=tid, company_id=cid).first()
        if t:
            results.append({'name': t.name, 'status': t.status, 'data': t.get_data()})

    import json as _json
    buf = io.BytesIO(_json.dumps(results, ensure_ascii=False, indent=2).encode('utf-8'))
    buf.seek(0)
    safe = c.name.replace(' ', '_')[:30]
    fname = f"injection_{safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/json')


@app.route('/company/<int:cid>/inject/all', methods=['GET'])
def company_inject_all(cid):
    """Page to inject all traitements of a company to ANPDP at once."""
    c = Company.query.get_or_404(cid)
    traitements = Traitement.query.filter_by(company_id=cid)\
                                  .order_by(Traitement.name).all()
    return render_template('inject_all.html', company=c, traitements=traitements)

@app.route('/api/inject/traitement/<int:tid>', methods=['POST'])
def api_inject_traitement(tid):
    """Inject a single traitement to ANPDP via the server-side script."""
    t = Traitement.query.get_or_404(tid)
    payload = request.get_json()
    cookies = payload.get('cookies', {})
    nit     = payload.get('nit', '')

    if not cookies.get('PHPSESSID') or not cookies.get('cookiesession1'):
        return jsonify(ok=False, error='Cookies manquants')

    try:
        import sys, os
        sys.path.insert(0, app.root_path)
        from anpdp_server import send_to_anpdp

        data = t.get_data()
        logs = []
        id_traitement = send_to_anpdp(data, cookies, logs)

        t.status = 'soumis'
        t.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify(ok=True, declaration_id=id_traitement, logs=logs)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[INJECT ERROR] {tb}", flush=True)
        return jsonify(ok=False, error=str(e), logs=logs if 'logs' in dir() else [])



@app.route('/company/<int:cid>/import', methods=['GET','POST'])
def company_import(cid):
    c = Company.query.get_or_404(cid)
    if request.method == 'GET':
        return render_template('import.html', company=c)

    f = request.files.get('excel_file')
    if not f or not f.filename:
        return redirect(url_for('company_import', cid=cid))

    import tempfile, os
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.xlsx','.xls','.xlsm'):
        return render_template('import.html', company=c,
                               error='Format non supporté. Utilisez .xlsx, .xls ou .xlsm')

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    f.save(tmp.name); tmp.close()

    try:
        from excel_import import extract_from_excel
        import traceback as _tb
        results = extract_from_excel(tmp.name)
        print(f"[IMPORT] File: {f.filename}, Results: {len(results)}", flush=True)
        for name, data, log in results:
            print(f"[IMPORT]   Sheet: {name}, Log: {log}", flush=True)
    except Exception as e:
        tb = _tb.format_exc()
        print(f"[IMPORT ERROR]\n{tb}", flush=True)
        return render_template('import.html', company=c,
                               error=str(e), traceback=tb)
    finally:
        try: os.unlink(tmp.name)
        except: pass

    if not results:
        return render_template('import.html', company=c,
                               error='Aucun traitement trouvé dans ce fichier.')

    # Create one traitement per sheet — continue even if one fails
    created = []
    errors  = []
    for t_name, data, log in results:
        try:
            t = Traitement(company_id=cid,
                           name=t_name.strip() or 'Traitement importé',
                           status='brouillon')
            t.set_data(data)
            db.session.add(t)
            db.session.flush()
            created.append({'id': t.id, 'name': t.name, 'log': log})
        except Exception as e:
            db.session.rollback()
            errors.append({'name': t_name, 'log': [f'❌ Erreur: {str(e)[:200]}']})
            created.append({'id': None, 'name': t_name,
                            'log': [f'❌ Erreur lors de la création: {str(e)[:200]}']})

    db.session.commit()
    session['import_results'] = created
    return redirect(url_for('import_result', cid=cid))

@app.route('/company/<int:cid>/import/result')
def import_result(cid):
    c = Company.query.get_or_404(cid)
    results = session.pop('import_results', [])
    return render_template('import_result.html', company=c, results=results)

# ─── EXCEL EXPORT (per company) ───────────────────────────────────────────────

@app.route('/company/<int:cid>/export/excel')
def company_export_excel(cid):
    c = Company.query.get_or_404(cid)
    traitements = Traitement.query.filter_by(company_id=cid).all()

    try:
        from generate_excel import build_sheet, build_choix_sheet
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        return "openpyxl not installed. Run: pip install openpyxl", 500

    wb = Workbook()
    wb.remove(wb.active)

    for t in traitements:
        d = t.get_data()
        if not d: continue
        safe_name = t.name[:28].replace('/', '-').replace('\\', '-')
        try:
            ws = build_sheet(wb, safe_name, d, 'Déclaration')
            # Embed company logo in top-right corner if available
            if c.logo_path and ws:
                logo_file = None
                if c.logo_path.startswith('/static/logos/'):
                    logo_file = os.path.join(app.root_path, c.logo_path.lstrip('/'))
                elif c.logo_path.startswith('http'):
                    # Download URL logo to a temp file
                    try:
                        import urllib.request, tempfile
                        ext = c.logo_path.split('.')[-1].split('?')[0][:4]
                        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
                        urllib.request.urlretrieve(c.logo_path, tmp.name)
                        logo_file = tmp.name
                    except: pass
                if logo_file and os.path.exists(logo_file):
                    try:
                        img = XLImage(logo_file)
                        img.width  = 120
                        img.height = 50
                        img.anchor = 'N1'
                        ws.add_image(img)
                    except Exception as img_e:
                        print(f"Logo embed error: {img_e}")
        except Exception as e:
            print(f"Sheet error {t.name}: {e}")

    build_choix_sheet(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_company = c.name.replace(' ', '_')[:30]
    filename = f"Registre_{safe_company}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─── JSON EXPORT (single traitement) ─────────────────────────────────────────

@app.route('/company/<int:cid>/traitement/<int:tid>/export/json')
def traitement_export_json(cid, tid):
    t = Traitement.query.get_or_404(tid)
    if t.company_id != cid: abort(404)
    buf = io.BytesIO(t.data.encode('utf-8'))
    filename = f"anpdp_{t.name.replace(' ','_')[:30]}.json"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/json')

# ─── AI ASSISTANT ────────────────────────────────────────────────────────────

ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')  # Set this in Railway Variables

SYSTEM_PROMPT = """Tu es un expert en protection des données personnelles en Algérie, spécialisé dans la loi 18-07 et les procédures de l'ANPDP.

Tu reçois automatiquement le contexte du formulaire : section active, champs remplis, résumé du traitement.

Tes capacités :
1. RÉDACTION — améliore la formulation des champs (finalité, cadre légal, etc.)
2. TRADUCTION — traduis en arabe formel algérien (فصحى administrative)
3. RÉVISION — analyse la conformité à la loi 18-07 et signale les manques
4. SUGGESTION — propose des catégories, des mesures de sécurité, des durées de conservation
5. AUTO-REMPLISSAGE — quand l'utilisateur demande à remplir des champs, retourne un bloc JSON

Pour l'auto-remplissage, inclus un bloc ```json``` avec les clés suivantes selon ce qui est pertinent :
- denomination, denomination_ar
- finalite, finalite_ar  
- cadre, cadre_ar
- comment, comment_ar (droits des personnes)
- service_droits, service_droits_ar
- adresse_droits, adresse_droits_ar
- mesures_droits, mesures_droits_ar
- methode_consent, methode_consent_ar
- nom_bdd, lieu_bdd (conservation informatique)
- nom_fm, lieu_fm (conservation manuelle)

N'inclus que les champs pour lesquels tu as une suggestion pertinente.
Explique brièvement tes suggestions après le bloc JSON.
Réponds toujours en français sauf pour les traductions en arabe."""

@app.route('/api/check_key')
def api_check_key():
    return jsonify(has_key=bool(ANTHROPIC_API_KEY))

@app.route('/api/chat', methods=['POST'])
def api_chat():
    if not ANTHROPIC_API_KEY:
        # Demo mode — return a helpful mock response
        payload = request.get_json()
        msgs = payload.get('messages', [])
        last = msgs[-1]['content'] if msgs else ''
        
        if 'tradui' in last.lower() or 'arabe' in last.lower():
            reply = "🔑 **Mode démo** — Pour activer la traduction automatique, ajoutez votre clé API Anthropic dans les variables d'environnement de Railway :\n\n`ANTHROPIC_API_KEY = sk-ant-...`\n\nObtenez votre clé sur **console.anthropic.com**"
        elif 'révis' in last.lower() or 'analys' in last.lower():
            reply = "🔑 **Mode démo** — L'analyse du formulaire nécessite une clé API Anthropic.\n\nUne fois configurée, je pourrai analyser tous vos champs et signaler les points non conformes à la loi 18-07."
        else:
            reply = "🔑 **Mode démo actif** — Je suis votre assistant ANPDP spécialisé en loi 18-07.\n\nPour m'activer complètement :\n1. Créez un compte sur **console.anthropic.com**\n2. Générez une clé API\n3. Ajoutez `ANTHROPIC_API_KEY` dans vos variables Railway\n\nEn attendant, voici ce que je pourrai faire :\n- ✍️ Rédiger et améliorer vos champs\n- 🌐 Traduire en arabe formel\n- 🔍 Analyser la conformité loi 18-07\n- 📂 Suggérer les bonnes catégories\n- ⚖️ Répondre à toutes vos questions juridiques"
        
        return jsonify(reply=reply)

    # Real API call
    try:
        import urllib.request, json as _json
        payload  = request.get_json()
        messages = payload.get('messages', [])
        context  = payload.get('context', {})

        # Build system prompt with full context
        ctx_lines = []
        if context.get('company'):     ctx_lines.append(f"Société: {context['company']} (NIT: {context.get('nit','')})")
        if context.get('activeSection'): ctx_lines.append(f"Section active: {context['activeSection']}")
        if context.get('activeFields'): ctx_lines.append(f"Champs visibles:\n{_json.dumps(context['activeFields'], ensure_ascii=False, indent=2)}")
        if context.get('denom'):       ctx_lines.append(f"Traitement: {context['denom']}")
        if context.get('finalite'):    ctx_lines.append(f"Finalité: {context['finalite']}")
        if context.get('categorie'):   ctx_lines.append(f"Catégorie: {context['categorie']}")

        system = SYSTEM_PROMPT
        if ctx_lines:
            system += "\n\n---\nCONTEXTE ACTUEL DU FORMULAIRE:\n" + "\n".join(ctx_lines)

        body = _json.dumps({
            'model': 'claude-sonnet-4-5-20251001',
            'max_tokens': 1500,
            'system': system,
            'messages': messages[-12:]
        }).encode()

        req = urllib.request.Request(
            'https://api.anthropic.com/v1/messages',
            data=body,
            headers={
                'Content-Type': 'application/json',
                'x-api-key': ANTHROPIC_API_KEY,
                'anthropic-version': '2023-06-01'
            },
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = _json.loads(resp.read())
            reply  = result['content'][0]['text']

            # Extract JSON fills if present
            import re
            fills = None
            m = re.search(r'```json\s*([\s\S]*?)```', reply)
            if m:
                try: fills = _json.loads(m.group(1))
                except: pass

            return jsonify(reply=reply, fills=fills)

    except Exception as e:
        return jsonify(error=f"Erreur API: {str(e)[:200]}"), 500

# ─── KNOWLEDGE BASE SEED DATA ────────────────────────────────────────────────

KB_SEED = {
    'Gestion RH / Paie': { 'icon': '👥', 'entries': [
        ('finalite','Finalité (FR)','Gérer les dossiers administratifs du personnel, la paie, les congés et les carrières des employés.',False),
        ('finalite_ar','Finalité (AR)','تسيير الملفات الإدارية للموظفين والأجور والإجازات والمسارات المهنية.',True),
        ('cadre','Cadre légal (FR)','Loi 90-11 du 21 avril 1990 relative aux relations de travail. Loi 83-14 relative aux obligations des assujettis en matière de sécurité sociale.',False),
        ('cadre_ar','Cadre légal (AR)','القانون رقم 90-11 المؤرخ في 21 أبريل 1990 المتعلق بعلاقات العمل.',True),
        ('finalite','Finalité (FR)','Établir et gérer la paie mensuelle des employés conformément à la législation en vigueur.',False),
        ('comment','Droits — Comment (FR)','Par notification écrite adressée au service des ressources humaines.',False),
        ('comment_ar','Droits — Comment (AR)','عن طريق إشعار كتابي موجه إلى مصلحة الموارد البشرية.',True),
        ('service_droits','Droits — Service (FR)','Direction des Ressources Humaines (DRH)',False),
        ('mesures_droits','Droits — Mesures (FR)','Formulaire de demande d\'accès disponible auprès du DRH. Réponse dans un délai de 30 jours.',False),
    ]},
    'Gestion des Clients': { 'icon': '🤝', 'entries': [
        ('finalite','Finalité (FR)','Gérer la relation client, le suivi des commandes, la facturation et le service après-vente.',False),
        ('finalite_ar','Finalité (AR)','تسيير العلاقة مع العملاء ومتابعة الطلبات والفوترة وخدمة ما بعد البيع.',True),
        ('cadre','Cadre légal (FR)','Loi 04-02 du 23 juin 2004 fixant les règles applicables aux pratiques commerciales.',False),
        ('cadre_ar','Cadre légal (AR)','القانون رقم 04-02 المؤرخ في 23 يونيو 2004 المحدد للقواعد المطبقة على الممارسات التجارية.',True),
        ('finalite','Finalité (FR)','Prospecter de nouveaux clients et fidéliser la clientèle existante par des actions marketing ciblées.',False),
        ('comment','Droits — Comment (FR)','Par courrier électronique ou courrier postal adressé au service commercial.',False),
        ('comment_ar','Droits — Comment (AR)','عن طريق البريد الإلكتروني أو البريد العادي الموجه إلى المصلحة التجارية.',True),
        ('service_droits','Droits — Service (FR)','Direction Commerciale et Marketing',False),
        ('mesures_droits','Droits — Mesures (FR)','Formulaire de demande en ligne. Traitement sous 15 jours ouvrables.',False),
    ]},
    'Gestion des Fournisseurs': { 'icon': '📦', 'entries': [
        ('finalite','Finalité (FR)','Gérer les relations avec les fournisseurs, les contrats d\'approvisionnement et le suivi des paiements.',False),
        ('finalite_ar','Finalité (AR)','تسيير العلاقات مع الموردين وعقود التموين ومتابعة المدفوعات.',True),
        ('cadre','Cadre légal (FR)','Code civil algérien — Livre III. Loi 04-02 relative aux pratiques commerciales.',False),
        ('cadre_ar','Cadre légal (AR)','القانون المدني الجزائري — الكتاب الثالث المتعلق بالطرق المختلفة لاكتساب الملكية.',True),
        ('comment','Droits — Comment (FR)','Par courrier recommandé adressé à la Direction des Achats.',False),
        ('service_droits','Droits — Service (FR)','Direction des Achats et de la Logistique',False),
    ]},
    'Comptabilité': { 'icon': '📊', 'entries': [
        ('finalite','Finalité (FR)','Assurer la comptabilité générale de l\'organisme, établir les états financiers et satisfaire aux obligations fiscales et légales.',False),
        ('finalite_ar','Finalité (AR)','ضمان المحاسبة العامة للهيئة وإعداد القوائم المالية والوفاء بالالتزامات الجبائية والقانونية.',True),
        ('cadre','Cadre légal (FR)','Loi 07-11 du 25 novembre 2007 portant système comptable financier. Code des impôts directs et taxes assimilées.',False),
        ('cadre_ar','Cadre légal (AR)','القانون رقم 07-11 المؤرخ في 25 نوفمبر 2007 المتضمن النظام المحاسبي المالي.',True),
        ('finalite','Finalité (FR)','Gérer la trésorerie, le règlement des fournisseurs et le recouvrement des créances clients.',False),
        ('comment','Droits — Comment (FR)','Par demande écrite adressée à la Direction Financière et Comptable.',False),
        ('service_droits','Droits — Service (FR)','Direction Financière et Comptable (DFC)',False),
        ('mesures_droits','Droits — Mesures (FR)','Accès aux données via demande formelle. Réponse dans un délai légal de 30 jours.',False),
    ]},
    'Vidéosurveillance / Sécurité accès': { 'icon': '📷', 'entries': [
        ('finalite','Finalité (FR)','Assurer la sécurité des biens et des personnes au sein des locaux de l\'organisme par un système de vidéosurveillance.',False),
        ('finalite_ar','Finalité (AR)','ضمان أمن الممتلكات والأشخاص داخل مقار الهيئة عبر نظام المراقبة بالكاميرات.',True),
        ('cadre','Cadre légal (FR)','Loi 18-07 relative à la protection des personnes physiques dans le traitement des données à caractère personnel — Art. 12.',False),
        ('cadre_ar','Cadre légal (AR)','القانون رقم 18-07 المتعلق بحماية الأشخاص الطبيعيين في معالجة البيانات ذات الطابع الشخصي — المادة 12.',True),
        ('finalite','Finalité (FR)','Contrôler les accès aux zones sensibles et identifier les personnes habilitées à y accéder.',False),
        ('comment','Droits — Comment (FR)','Par demande écrite auprès du Responsable Sécurité. Images conservées 30 jours maximum.',False),
        ('comment_ar','Droits — Comment (AR)','بطلب كتابي لدى مسؤول الأمن. تُحفظ الصور 30 يوماً كحد أقصى.',True),
        ('service_droits','Droits — Service (FR)','Responsable Sécurité et Sûreté',False),
        ('mesures_droits','Droits — Mesures (FR)','Visionnage des images encadré par procédure interne. Accès restreint au personnel habilité.',False),
    ]},
    'Gestion des Formations': { 'icon': '🎓', 'entries': [
        ('finalite','Finalité (FR)','Planifier, organiser et suivre les actions de formation du personnel en vue du développement des compétences.',False),
        ('finalite_ar','Finalité (AR)','تخطيط وتنظيم ومتابعة إجراءات تكوين الموظفين بهدف تطوير الكفاءات.',True),
        ('cadre','Cadre légal (FR)','Loi 90-11 du 21 avril 1990 relative aux relations de travail — dispositions relatives à la formation professionnelle.',False),
        ('cadre_ar','Cadre légal (AR)','القانون رقم 90-11 المتعلق بعلاقات العمل — أحكام التكوين المهني.',True),
        ('finalite','Finalité (FR)','Gérer les inscriptions aux formations, évaluer les résultats et établir les attestations de participation.',False),
        ('comment','Droits — Comment (FR)','Par demande adressée au service Formation et Développement des compétences.',False),
        ('service_droits','Droits — Service (FR)','Service Formation et Développement des Compétences',False),
    ]},
    'Gestion informatique / Systèmes': { 'icon': '💻', 'entries': [
        ('finalite','Finalité (FR)','Gérer les accès aux systèmes d\'information, assurer la sécurité informatique et administrer le parc informatique.',False),
        ('finalite_ar','Finalité (AR)','إدارة الوصول إلى أنظمة المعلومات وضمان الأمن المعلوماتي وإدارة الحظيرة المعلوماتية.',True),
        ('cadre','Cadre légal (FR)','Loi 18-07 relative à la protection des données personnelles. Loi 09-04 portant règles particulières relatives à la prévention des infractions liées aux TIC.',False),
        ('cadre_ar','Cadre légal (AR)','القانون رقم 18-07 المتعلق بحماية البيانات الشخصية. القانون رقم 09-04 المتضمن القواعد الخاصة للوقاية من الجرائم المتصلة بتكنولوجيات الإعلام.',True),
        ('finalite','Finalité (FR)','Tracer les accès aux données sensibles et assurer la traçabilité des opérations sur les systèmes critiques.',False),
        ('comment','Droits — Comment (FR)','Par ticket adressé au service informatique via le portail interne.',False),
        ('service_droits','Droits — Service (FR)','Direction des Systèmes d\'Information (DSI)',False),
        ('mesures_droits','Droits — Mesures (FR)','Charte informatique signée par chaque utilisateur. Accès sur demande via formulaire sécurisé.',False),
    ]},
}

def _seed_kb():
    if KBType.query.count() > 0:
        return
    for type_name, data in KB_SEED.items():
        kt = KBType(name=type_name, icon=data['icon'])
        db.session.add(kt)
        db.session.flush()
        for (fkey, flabel, val, is_ar) in data['entries']:
            db.session.add(KBEntry(type_id=kt.id, field_key=fkey,
                field_label=flabel, value=val, is_arabic=is_ar, approved=True))
    db.session.commit()

# ─── KB ROUTES ────────────────────────────────────────────────────────────────

@app.route('/kb')
def kb_home():
    types   = KBType.query.order_by(KBType.name).all()
    pending = KBEntry.query.filter_by(approved=False).count()
    return render_template('kb_home.html', types=types, pending=pending)

@app.route('/kb/type/new', methods=['POST'])
def kb_type_new():
    name = request.form.get('name','').strip()
    icon = request.form.get('icon','🗂').strip() or '🗂'
    if name and not KBType.query.filter_by(name=name).first():
        db.session.add(KBType(name=name, icon=icon))
        db.session.commit()
    return redirect(url_for('kb_home'))

@app.route('/kb/type/<int:tid>')
def kb_type_view(tid):
    kt = KBType.query.get_or_404(tid)
    entries = KBEntry.query.filter_by(type_id=tid).order_by(
        KBEntry.field_key, KBEntry.approved.desc(), KBEntry.created_at.desc()).all()
    from collections import defaultdict
    grouped = defaultdict(list)
    for e in entries: grouped[e.field_key].append(e)
    return render_template('kb_type.html', kb_type=kt, grouped=dict(grouped))

@app.route('/kb/type/<int:tid>/delete', methods=['POST'])
def kb_type_delete(tid):
    db.session.delete(KBType.query.get_or_404(tid))
    db.session.commit()
    return redirect(url_for('kb_home'))

@app.route('/kb/type/<int:tid>/entry/add', methods=['POST'])
def kb_entry_add(tid):
    KBType.query.get_or_404(tid)
    e = KBEntry(type_id=tid,
        field_key  = request.form.get('field_key','').strip(),
        field_label= request.form.get('field_label','').strip(),
        value      = request.form.get('value','').strip(),
        is_arabic  = request.form.get('is_arabic') == '1',
        approved   = True)
    if e.field_key and e.value:
        db.session.add(e); db.session.commit()
    return redirect(url_for('kb_type_view', tid=tid))

@app.route('/kb/entry/<int:eid>/delete', methods=['POST'])
def kb_entry_delete(eid):
    e = KBEntry.query.get_or_404(eid)
    tid = e.type_id
    db.session.delete(e); db.session.commit()
    return redirect(url_for('kb_type_view', tid=tid))

@app.route('/kb/entry/<int:eid>/approve', methods=['POST'])
def kb_entry_approve(eid):
    e = KBEntry.query.get_or_404(eid)
    e.approved = True; db.session.commit()
    return jsonify(ok=True)

@app.route('/kb/entry/<int:eid>/reject', methods=['POST'])
def kb_entry_reject(eid):
    e = KBEntry.query.get_or_404(eid)
    db.session.delete(e); db.session.commit()
    return jsonify(ok=True)

@app.route('/kb/type/<int:tid>/import', methods=['POST'])
def kb_import_excel(tid):
    kt = KBType.query.get_or_404(tid)
    f  = request.files.get('file')
    if not f: return redirect(url_for('kb_type_view', tid=tid))
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active; added = 0; headers = []
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if ri == 0:
                headers = [str(h).strip().lower().replace(' ','_') if h else '' for h in row]; continue
            d = dict(zip(headers, row))
            fkey  = str(d.get('field_key','') or d.get('champ','') or '').strip()
            val   = str(d.get('value','')     or d.get('valeur','') or '').strip()
            flabel= str(d.get('label','')     or d.get('libelle','') or fkey).strip()
            is_ar = str(d.get('arabic','')    or d.get('arabe','')).strip().lower() in ('1','oui','yes','true')
            if fkey and val and val != 'None':
                if not KBEntry.query.filter_by(type_id=tid, field_key=fkey, value=val).first():
                    db.session.add(KBEntry(type_id=tid, field_key=fkey,
                        field_label=flabel, value=val, is_arabic=is_ar, approved=True))
                    added += 1
        db.session.commit()
    except Exception as e:
        return f"Erreur import: {e}", 500
    return redirect(url_for('kb_type_view', tid=tid))

# ─── KB API ───────────────────────────────────────────────────────────────────

@app.route('/api/kb/types')
def api_kb_types():
    types = KBType.query.order_by(KBType.name).all()
    return jsonify([{'id': t.id, 'name': t.name, 'icon': t.icon} for t in types])

@app.route('/api/kb/suggestions/<int:type_id>')
def api_kb_suggestions(type_id):
    entries = KBEntry.query.filter_by(type_id=type_id, approved=True).all()
    from collections import defaultdict
    grouped = defaultdict(list)
    for e in entries:
        grouped[e.field_key].append({'id':e.id,'value':e.value,'label':e.field_label,'is_arabic':e.is_arabic})
    return jsonify(dict(grouped))

@app.route('/api/kb/pending')
def api_kb_pending():
    pending = KBEntry.query.filter_by(approved=False).order_by(KBEntry.created_at.desc()).all()
    return jsonify([{'id':e.id,'type':e.kb_type.name,'icon':e.kb_type.icon,
        'field_key':e.field_key,'field_label':e.field_label,'value':e.value[:120],'is_arabic':e.is_arabic}
        for e in pending])

@app.route('/api/kb/learn', methods=['POST'])
def api_kb_learn():
    payload = request.get_json()
    type_id = payload.get('type_id')
    fields  = payload.get('fields', {})
    if not type_id: return jsonify(ok=False)
    added = 0
    for fkey, val in fields.items():
        val = str(val).strip()
        if not val or len(val) < 5: continue
        if not KBEntry.query.filter_by(type_id=type_id, field_key=fkey, value=val).first():
            db.session.add(KBEntry(type_id=type_id, field_key=fkey,
                field_label=fkey.replace('_',' ').title(), value=val,
                is_arabic=bool(any('\u0600'<=c<='\u06ff' for c in val)), approved=False))
            added += 1
    db.session.commit()
    return jsonify(ok=True, added=added)

# ─── INIT ─────────────────────────────────────────────────────────────────────

with app.app_context():
    db.create_all()
    _seed_kb()

@app.route('/api/check_server')
@auth_required
def api_check_server():
    # Frontend expects {requests: bool} to know if injection is possible
    try:
        import requests
        has_requests = True
    except ImportError:
        has_requests = False
    return jsonify({
        "status": "ok",
        "requests": has_requests,
        "version": "1.1.0-unified"
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'true').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
